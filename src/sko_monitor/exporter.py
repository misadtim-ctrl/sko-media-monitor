from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADERS = [
    "Дата",
    "Время",
    "Источник",
    "Название СМИ",
    "Заголовок",
    "Ссылка",
    "Тип источника",
    "Краткое содержание",
    "Тональность",
    "Релевантность",
    "Ключевые темы",
]


def _rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    result: list[list[Any]] = []
    for item in items:
        publication = item["publication"]
        analysis = item["analysis"]
        published = publication.get("published_at") or ""
        date, time = "", ""
        if "T" in published:
            date, time = published.split("T", 1)
            time = time[:8]
        result.append(
            [
                date,
                time,
                publication.get("source_name", ""),
                publication.get("source_name", ""),
                publication.get("title", ""),
                publication.get("url", ""),
                publication.get("platform", ""),
                analysis.get("summary", ""),
                analysis.get("tone", ""),
                analysis.get("confidence", 0),
                ", ".join(analysis.get("matched", [])),
            ]
        )
    return result


def export_latest(items: list[dict[str, Any]], export_dir: Path) -> tuple[Path, Path]:
    export_dir.mkdir(parents=True, exist_ok=True)
    csv_path = export_dir / "latest.csv"
    xlsx_path = export_dir / "latest.xlsx"
    rows = _rows(items)

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Мониторинг"
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    widths = [13, 10, 20, 20, 55, 45, 18, 55, 16, 16, 35]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(xlsx_path)
    return csv_path, xlsx_path


def export_historical(items: list[dict[str, Any]], export_dir: Path) -> tuple[Path, Path]:
    export_dir.mkdir(parents=True, exist_ok=True)
    csv_path = export_dir / "historical-latest.csv"
    xlsx_path = export_dir / "historical-latest.xlsx"
    headers = ["Дата", "Источник", "Заголовок", "Ссылка", "Кратко", "Релевантность", "Метод"]
    rows = [
        [
            item.get("published_at", "")[:10],
            item.get("source", ""),
            item.get("title", ""),
            item.get("url", ""),
            item.get("summary", ""),
            item.get("relevance", 0),
            item.get("method", ""),
        ]
        for item in items
    ]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Поиск за период"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    header_fill = PatternFill("solid", fgColor="4527A0")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    widths = [14, 22, 55, 45, 55, 16, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(xlsx_path)
    return csv_path, xlsx_path
